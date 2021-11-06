# -*- coding: utf-8 -*-
"""
Created on Thu Jan 28 15:20:53 2021

This script automates the pulling of Purolator tracking details such as ship to, ship from and status of delivery.

@author: Kenneth Liang
"""
# Open Write Excel files
from openpyxl import load_workbook
from xlsxwriter import Workbook

# Chrome browser control
from selenium import webdriver
from selenium.webdriver import ChromeOptions 

# For directory
from os import getcwd

# For GUI
import PySimpleGUI as sim_gui

# Initialize variable to store the purolator details scraped from site
tracking_details = []
link = []
no_work_book = True

# create browser window, this needs to use the chrome driver exe
options = ChromeOptions() 
options.headless = True
browser = webdriver.Chrome(executable_path=getcwd()+'\\chromedriver.exe',options=options)
browser.implicitly_wait(3.5)

# open excel file that holds all the purolator tracking numbers that need to be looked up
try:
    workbook = load_workbook(getcwd()+'\\Tracking_Number.xlsx')
    
    no_work_book = False
    
except:
    sim_gui.popup('Did not find file Tracking_Number.xlsx script will now close')

if no_work_book == False:
    
    # Initialize workbook variables
    sheet = workbook.active
    max_r = sheet.max_row
    
    # Initialize gui progress bar window
    # layout the window
    layout = [[sim_gui.Text('Please wait as script runs')],
              [sim_gui.ProgressBar(max_r, orientation='h', size=(20, 20), key='progressbar')],
              [sim_gui.Cancel()]]
    # create the window
    window = sim_gui.Window('Progress Meter', layout)
    progress_bar = window['progressbar']
    
    for row in range(2,max_r+1):
        
        # check to see if the cancel button was clicked and exit loop if clicked
        event, values = window.read(timeout=10)
        
        if event == 'Cancel' or event == sim_gui.WIN_CLOSED:
            break
        
        else:    
            tracking_num = int(sheet.cell(row,1).value)

            browser.get(f'https://www.purolator.com/en/shipping/tracker?pin={tracking_num}')
            S = lambda X: browser.execute_script('return document.body.parentNode.scroll'+X)
            browser.set_window_size(S('Width'), S('Height'))

            try:
                status = browser.find_element_by_class_name('col-12.col-sm-7').text
            except:
                status = 'Was not able to find status'

            # There are multiple of the following element and find_elements returns a list of these elements
            try:
                element_list = browser.find_elements_by_class_name('col-7.col-sm-12.col-md-7')

                origin = element_list[0].text
                destination = element_list[1].text
                reference = element_list[2].text
                service = element_list[3].text
                weight = element_list[4].text
                ship_date = element_list[5].text

            except:
                origin = 'No origin'
                destination = 'No destination'
                reference = 'No reference'
                service = 'No service'
                weight = 'No weight'
                ship_date = 'No ship date'

            try:
                history = browser.find_element_by_class_name('history').text
            except:
                history = 'Was not able to find history'

            try:
                browser.find_element_by_class_name('results-contents.container-fluid').screenshot(getcwd()+f'\\Screen_Shots\\tracking_details{row-1}.png')
                screen_shot = getcwd()+f'\\Screen_Shots\\tracking_details{row-1}.png'
            except:
                screen_shot = 'No screen shot taken'

            # Append results of web scraping to a list
            tracking_details.append([tracking_num,status,origin,destination,reference,service,weight,ship_date,history])
            link.append(screen_shot)
            
            # update bar with loop value +1 so that bar eventually reaches the maximum
            progress_bar.UpdateBar(row-1)

    # Close chrome browser and workbook
    browser.close()
    browser.quit()
    workbook.close()

    # Output results of web scraping to a new excel file
    with Workbook('Tracking_Details.xlsx') as newworkbook:
        worksheet = newworkbook.add_worksheet()

        # First row are the header names
        worksheet.write_row(0,0,['Tracking Number','Status','Origin','Destination','Reference','Service','Weight','Ship Date','History','Screen Shot Link'])

        # Output the row data
        for row_num, data in enumerate(tracking_details):
            worksheet.write_row(row_num+1, 0, data, newworkbook.add_format({'text_wrap':True, 'valign':'vcenter'}))

        # Output the screen shot data
        for row_num, data in enumerate(link):
            worksheet.write_url(row=row_num+1, col=9, url='external:'+data, string='Link', cell_format=newworkbook.add_format({'text_wrap':True, 'valign':'vcenter'}))

        worksheet.set_column(worksheet.dim_colmax-1,worksheet.dim_colmax-1,80)
    
    window.close()